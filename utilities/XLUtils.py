import openpyxl
from openpyxl.styles import PatternFill
def get_row_count(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    return worksheet.max_row

def get_column_count(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    return worksheet.max_column

def read_excel(file_name, sheet_name, row, col):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    return worksheet.cell(row=row, column=col).value

def write_to_excel(file_name, sheet_name, data, row, col):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    worksheet.cell(row=row, column=col). value = data
    workbook.save(file_name)

def create_excel_workbook(file_name):
    workbook = openpyxl.Workbook()
    workbook.save(file_name)

def fill_green(file_name, sheet_name, row, col):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    fill = PatternFill(start_color='60b212', end_color='60b212', fill_type="solid")
    cell = worksheet.cell(row=row, column=col)
    cell.fill = fill
    workbook.save(file_name)

def fill_red(file_name, sheet_name, row, col):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")
    cell = worksheet.cell(row=row, column=col)
    cell.fill = fill
    workbook.save(file_name)

"""create_excel_workbook("data.xlsx")

for r in range(1,11):
    for c in range(1,6):
        write_to_excel("data.xlsx","Sheet","hi", r, c)
        print(read_excel("data.xlsx","Sheet",r,c))
    print()

row=get_row_count("data.xlsx","Sheet")
col=get_column_count("data.xlsx","Sheet")
print(row,col)

fill_green("data.xlsx","Sheet",1,5)
fill_red("data.xlsx","Sheet",5,5)
"""